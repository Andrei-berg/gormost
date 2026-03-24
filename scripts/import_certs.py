#!/usr/bin/env python3
"""
Import employee certificates from Excel into Supabase.
Usage:
  pip install openpyxl requests python-dotenv
  python3 scripts/import_certs.py [--dry-run] [--sheet <name>]

Options:
  --dry-run   Print what would be inserted, don't write to DB
  --sheet     Import only the specified sheet (default: all)
"""

from __future__ import annotations
import sys
import os
import re
import json
import argparse
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

try:
    import openpyxl
    import requests
    from dotenv import load_dotenv
except ImportError as e:
    print(f"Missing dependency: {e}")
    print("Run: pip install openpyxl requests python-dotenv")
    sys.exit(1)

# ============================================================
# CONFIG
# ============================================================

EXCEL_PATH = Path(__file__).parent.parent / "downloads" / "1. моя УДОСТОВЕРЕНИЯ ПО КАТЕГОРИЯМ.xlsx"

# Try to find Excel in common locations
EXCEL_CANDIDATES = [
    EXCEL_PATH,
    Path.home() / "Downloads" / "1. моя УДОСТОВЕРЕНИЯ ПО КАТЕГОРИЯМ.xlsx",
    Path(__file__).parent.parent / "1. моя УДОСТОВЕРЕНИЯ ПО КАТЕГОРИЯМ.xlsx",
]

# Map sheet name → cert_type code
SHEET_TO_CERT_TYPE = {
    'Бригадиры':             'BRIGADIER',
    'Стропальщик':           'SLING',
    'Высота с подмащ':       'HEIGHT_SCAFFOLD',
    'Высота 1 гр ':          'HEIGHT_1',
    'Высота 2 гр':           'HEIGHT_2',
    'Высота 3 гр':           'HEIGHT_3',
    'Люлька':                'GONDOLA',
    'Пескоструйщик':         'SANDBLAST',
    'Маш Компрессорн уст-ки':'COMPRESSOR_OP',
    'Маш Фасадн подъемников':'FACADE_LIFT_OP',
    'Маш Крана-манипулятора':'MANIPULATOR_OP',
    'Маш подъемника (вышки)':'LIFT',
    'Электрогазосварщик':    'NAKS',
    'Изолировщик':           'INSULATOR',
    'Земляные работы':       'EARTHWORKS',
    'осн туж ландш пож':     'PTM',
    # Special: ОТ and Промбез handled separately (multiple certs per row)
}

# ============================================================
# HELPERS
# ============================================================

def normalize_name(name: str) -> str:
    """Normalize full name for fuzzy matching."""
    if not name:
        return ''
    # Strip whitespace, collapse multiple spaces, remove notes in parens
    n = re.sub(r'\s+', ' ', str(name).strip())
    # Remove medical notes like 'МЕДОТВОД'
    n = re.sub(r'\s+МЕДОТВОД.*$', '', n, flags=re.IGNORECASE)
    return n.strip()


def parse_date(val) -> str | None:
    """Parse various date formats → 'YYYY-MM-DD' or None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, str):
        # Formula like =F2+365 — can't resolve without context
        if val.startswith('='):
            return None
        # "бессрочно"
        if 'бессрочно' in val.lower():
            return None
        # Try to parse date strings
        for fmt in ('%d.%m.%Y', '%d.%m.%y', '%Y-%m-%d'):
            try:
                return datetime.strptime(val.strip(), fmt).strftime('%Y-%m-%d')
            except ValueError:
                continue
    if isinstance(val, (int, float)):
        # Excel serial date number
        try:
            origin = datetime(1899, 12, 30)
            return (origin + timedelta(days=int(val))).strftime('%Y-%m-%d')
        except Exception:
            return None
    return None


def calc_expires_from_formula(formula: str, issued_at: str | None) -> str | None:
    """Calculate expiry date from Excel formula like =F2+365."""
    if not issued_at or not formula:
        return None
    m = re.search(r'\+(\d+)$', formula)
    if m:
        days = int(m.group(1))
        try:
            issued = datetime.strptime(issued_at, '%Y-%m-%d')
            return (issued + timedelta(days=days)).strftime('%Y-%m-%d')
        except Exception:
            return None
    return None


def parse_expires(val, issued_at: str | None) -> tuple[str | None, bool]:
    """Returns (expires_at, is_indefinite)."""
    if val is None:
        return (None, False)
    if isinstance(val, str):
        lower = val.lower().strip()
        if 'бессрочно' in lower:
            return (None, True)
        if lower.startswith('='):
            calculated = calc_expires_from_formula(lower, issued_at)
            return (calculated, False)
    date_str = parse_date(val)
    return (date_str, False)


def parse_ob_pa_date(val) -> str | None:
    """Parse 'ОБ 10.04.26', 'ПА 24.04.26', or datetime."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, str):
        # Remove prefix like 'ОБ ', 'ПА ', 'обучение ' etc.
        cleaned = re.sub(r'^[А-Яа-я/]+\s+', '', val.strip())
        for fmt in ('%d.%m.%Y', '%d.%m.%y', '%d.%m.%Y '):
            try:
                return datetime.strptime(cleaned.strip(), fmt).strftime('%Y-%m-%d')
            except ValueError:
                continue
    return None


def parse_renewal_type(val) -> str | None:
    if not val:
        return None
    s = str(val).lower().strip()
    if 'первичное' in s:
        return 'первичное'
    if 'очередное' in s:
        return 'очередное'
    if 'повторное' in s:
        return 'повторное'
    if 'бессрочное' in s:
        return 'бессрочное'
    return None


def parse_return_status(val) -> str | None:
    if not val:
        return None
    s = str(val).lower()
    if 'забрал' in s:
        return 'забрал'
    if 'не сдал' in s:
        return 'не сдал'
    if 'сдал' in s:
        return 'сдал'
    return None


def parse_shift(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    # Normalize various shift notations
    s = re.sub(r'\\', '/', s)  # 6\6 → 6/6
    s = re.sub(r'^вахта\s+', '', s, flags=re.IGNORECASE)
    s = s.strip()
    return s if s else None


# ============================================================
# SHEET PARSERS
# ============================================================

def parse_standard_sheet(ws, cert_type_code: str) -> list[dict]:
    """
    Parse standard cert sheet with columns:
    смена | ФИО | Профессия | Обучение | № уд | Дата обучения | Действительно до |
    № Протокола | первичное/очередное | Дата ОБ/ПА | Комментарий | сдал/не сдал

    Some sheets have extra leading column (№ п/п), detected automatically.
    """
    records = []
    rows = list(ws.iter_rows(values_only=True))

    # Find header row
    header_row_idx = None
    for i, row in enumerate(rows):
        if row and any(v and 'ФИО' in str(v) for v in row):
            header_row_idx = i
            break
    if header_row_idx is None:
        return records

    headers = [str(h).strip().lower() if h else '' for h in rows[header_row_idx]]

    # Determine column offsets based on header
    def col(keyword):
        for i, h in enumerate(headers):
            if keyword.lower() in h:
                return i
        return None

    i_shift = col('смен') or col('№ смен') or 0
    i_name = col('фио') or col('сотрудник') or col('ф.и.о')
    i_doc = col('удо') or col('№ уд')
    i_issued = col('дата обучения') or col('дата выдачи')
    i_expires = col('действительно')
    i_protocol = col('протакола') or col('протокола')
    i_renewal = col('первичное')
    i_ob_pa = col('дата об') or col('об/па')
    i_notes = col('коммент') or col('примечани')
    i_return = col('сдал')

    for row in rows[header_row_idx + 1:]:
        if not row or not any(v is not None for v in row):
            continue
        # Skip section headers (service name rows have all nulls except 2nd col)
        name_val = row[i_name] if i_name is not None and i_name < len(row) else None
        if name_val is None:
            continue
        name_str = normalize_name(str(name_val))
        if not name_str or len(name_str) < 3:
            continue
        # Skip rows that look like sub-headers
        if any(word in name_str.lower() for word in ['служба', 'отдел', '№ п/п', 'фио', 'ф.и.о']):
            continue

        shift = parse_shift(row[i_shift] if i_shift is not None and i_shift < len(row) else None)
        doc_number = str(row[i_doc]).strip() if (i_doc is not None and i_doc < len(row) and row[i_doc] is not None) else None
        issued_raw = row[i_issued] if i_issued is not None and i_issued < len(row) else None
        expires_raw = row[i_expires] if i_expires is not None and i_expires < len(row) else None
        protocol = str(row[i_protocol]).strip() if (i_protocol is not None and i_protocol < len(row) and row[i_protocol] is not None) else None
        renewal = parse_renewal_type(row[i_renewal] if i_renewal is not None and i_renewal < len(row) else None)
        ob_pa = parse_ob_pa_date(row[i_ob_pa] if i_ob_pa is not None and i_ob_pa < len(row) else None)
        notes_val = row[i_notes] if i_notes is not None and i_notes < len(row) else None
        return_val = row[i_return] if i_return is not None and i_return < len(row) else None

        issued_at = parse_date(issued_raw)
        expires_at, is_indefinite = parse_expires(expires_raw, issued_at)

        # If renewal_type is 'бессрочное', mark indefinite
        if renewal == 'бессрочное':
            is_indefinite = True

        records.append({
            'full_name': name_str,
            'cert_type_code': cert_type_code,
            'shift_number': shift,
            'doc_number': doc_number,
            'issued_at': issued_at,
            'expires_at': expires_at,
            'is_indefinite': is_indefinite,
            'protocol_number': protocol,
            'renewal_type': renewal,
            'ob_pa_date': ob_pa,
            'notes': str(notes_val).strip() if notes_val else None,
            'return_status': parse_return_status(return_val),
        })

    return records


def parse_brigadiry_sheet(ws) -> list[dict]:
    """
    Бригадиры sheet:
    смена | ФИО | Профессия | № уд | № Протокола | Дата протокола | Действительно до |
    Дата выдачи | первичное/очередное | Дата ОБ/ПА | Комментарий | сдал/забрал
    """
    records = []
    rows = list(ws.iter_rows(values_only=True))

    for row in rows[2:]:  # Skip title + header
        if not row or not any(v is not None for v in row):
            continue
        name_val = row[1] if len(row) > 1 else None
        if name_val is None:
            continue
        name_str = normalize_name(str(name_val))
        if not name_str or len(name_str) < 3:
            continue
        if any(w in name_str.lower() for w in ['служба', 'отдел']):
            continue

        shift = parse_shift(row[0] if row[0] is not None else None)
        doc_number = str(row[3]).strip() if (len(row) > 3 and row[3] is not None) else None
        protocol = str(row[4]).strip() if (len(row) > 4 and row[4] is not None) else None
        # Дата протокола = col 5, Действительно до = col 6, Дата выдачи = col 7
        issued_at = parse_date(row[7] if len(row) > 7 else None)
        expires_raw = row[6] if len(row) > 6 else None
        expires_at, is_indefinite = parse_expires(expires_raw, issued_at)
        renewal = parse_renewal_type(row[8] if len(row) > 8 else None)
        ob_pa = parse_ob_pa_date(row[9] if len(row) > 9 else None)
        notes_val = row[10] if len(row) > 10 else None
        return_val = row[11] if len(row) > 11 else None

        records.append({
            'full_name': name_str,
            'cert_type_code': 'BRIGADIER',
            'shift_number': shift,
            'doc_number': doc_number,
            'issued_at': issued_at,
            'expires_at': expires_at,
            'is_indefinite': is_indefinite,
            'protocol_number': protocol,
            'renewal_type': renewal,
            'ob_pa_date': ob_pa,
            'notes': str(notes_val).strip() if notes_val else None,
            'return_status': parse_return_status(return_val),
        })

    return records


def parse_ot_sheet(ws) -> list[dict]:
    """
    ОТ sheet has one row per employee with up to 4 cert columns:
    Сотрудник | ... | Дата А | № А | Дата Б | № Б | Дата ПП | № ПП | Дата СИЗ | № СИЗ
    Produces up to 4 records per employee.
    """
    records = []
    rows = list(ws.iter_rows(values_only=True))

    # Header is row 0
    headers = [str(h).lower().strip() if h else '' for h in (rows[0] if rows else [])]

    # Column indices
    def col(keyword):
        for i, h in enumerate(headers):
            if keyword in h:
                return i
        return None

    i_name = col('сотрудник') or 2
    i_date_a = col('программе а')
    i_num_a = None
    i_date_b = col('программе б')
    i_num_b = None
    i_date_pp = col('первой помощи')
    i_num_pp = None
    i_date_siz = col('сиз')
    i_num_siz = None

    # Find number columns right after date columns
    def next_col(i):
        return i + 1 if i is not None else None

    i_num_a = next_col(i_date_a)
    i_num_b = next_col(i_date_b)
    i_num_pp = next_col(i_date_pp)
    i_num_siz = next_col(i_date_siz)

    cert_map = [
        ('OT_A',    i_date_a,  i_num_a),
        ('OT_B',    i_date_b,  i_num_b),
        ('FIRST_AID', i_date_pp, i_num_pp),
        ('OT_SIZ',  i_date_siz, i_num_siz),
    ]

    for row in rows[1:]:
        if not row or not any(v is not None for v in row):
            continue
        name_val = row[i_name] if i_name < len(row) else None
        if name_val is None:
            continue
        name_str = normalize_name(str(name_val))
        if not name_str or len(name_str) < 3:
            continue

        for cert_code, i_date, i_num in cert_map:
            if i_date is None:
                continue
            date_val = row[i_date] if i_date < len(row) else None
            num_val = row[i_num] if (i_num is not None and i_num < len(row)) else None
            if date_val is None and num_val is None:
                continue
            issued_at = parse_date(date_val)
            # OT certs: 3 year period (36 months)
            expires_at = None
            if issued_at:
                try:
                    issued = datetime.strptime(issued_at, '%Y-%m-%d')
                    expires_at = (issued + timedelta(days=1095)).strftime('%Y-%m-%d')
                except Exception:
                    pass
            doc_number = str(num_val).strip() if num_val else None

            records.append({
                'full_name': name_str,
                'cert_type_code': cert_code,
                'shift_number': None,
                'doc_number': doc_number,
                'issued_at': issued_at,
                'expires_at': expires_at,
                'is_indefinite': False,
                'protocol_number': None,
                'renewal_type': 'очередное',
                'ob_pa_date': None,
                'notes': None,
                'return_status': None,
            })

    return records


def parse_promsbez_sheet(ws) -> list[dict]:
    """
    Промбез sheet:
    смена | ФИО | Профессия | Обучение (description) | Статья | Дата обучения |
    Действительно до | Экзамен | № Протокола | Примечание

    Each row is a separate cert (one person can have A1 and B9.3 etc.)
    Map Статья: starts with 'А' → PROM_A1, starts with 'Б' → PROM_B
    """
    records = []
    rows = list(ws.iter_rows(values_only=True))

    for row in rows[1:]:  # skip header
        if not row or not any(v is not None for v in row):
            continue
        name_val = row[1] if len(row) > 1 else None
        if name_val is None:
            continue
        name_str = normalize_name(str(name_val))
        if not name_str or len(name_str) < 3:
            continue

        statya = str(row[4]).strip().upper() if (len(row) > 4 and row[4]) else ''
        cert_code = 'PROM_A1' if statya.startswith('А') or statya.startswith('A') else 'PROM_B'

        issued_at = parse_date(row[5] if len(row) > 5 else None)
        expires_raw = row[6] if len(row) > 6 else None
        expires_at, is_indefinite = parse_expires(expires_raw, issued_at)
        protocol = str(row[8]).strip() if (len(row) > 8 and row[8]) else None
        notes_val = row[9] if len(row) > 9 else None

        records.append({
            'full_name': name_str,
            'cert_type_code': cert_code,
            'shift_number': parse_shift(row[0]),
            'doc_number': None,
            'issued_at': issued_at,
            'expires_at': expires_at,
            'is_indefinite': is_indefinite,
            'protocol_number': protocol,
            'renewal_type': 'очередное',
            'ob_pa_date': None,
            'notes': f"Статья: {statya}" + (f". {notes_val}" if notes_val else ""),
            'return_status': None,
        })

    return records


# ============================================================
# MATCHING
# ============================================================

def build_name_index(users: list[dict]) -> dict[str, str]:
    """Build normalized_name → user_id index."""
    idx = {}
    for u in users:
        n = normalize_name(u.get('full_name', ''))
        if n:
            idx[n.lower()] = u['user_id']
    return idx


def fuzzy_match(name: str, name_idx: dict) -> str | None:
    """Try to find user_id by name. Returns user_id or None."""
    from difflib import get_close_matches
    normalized = name.lower().strip()
    if normalized in name_idx:
        return name_idx[normalized]
    # Try partial match (first 2 words)
    parts = normalized.split()
    if len(parts) >= 2:
        prefix = ' '.join(parts[:2])
        for idx_name, uid in name_idx.items():
            if idx_name.startswith(prefix):
                return uid
    # Fuzzy
    candidates = get_close_matches(normalized, name_idx.keys(), n=1, cutoff=0.85)
    if candidates:
        return name_idx[candidates[0]]
    return None


# ============================================================
# SUPABASE API
# ============================================================

def get_supabase_client():
    load_dotenv(Path(__file__).parent.parent / '.env.local')
    url = os.environ.get('NEXT_PUBLIC_SUPABASE_URL')
    key = os.environ.get('NEXT_PUBLIC_SUPABASE_ANON_KEY')
    if not url or not key:
        raise RuntimeError("Missing SUPABASE_URL or ANON_KEY in .env.local")
    return url, key


def fetch_users(url: str, key: str) -> list[dict]:
    headers = {'apikey': key, 'Authorization': f'Bearer {key}'}
    r = requests.get(f'{url}/rest/v1/users?select=user_id,full_name&is_active=eq.true&limit=500', headers=headers)
    r.raise_for_status()
    return r.json()


def fetch_cert_types(url: str, key: str) -> dict[str, str]:
    """Returns code → id map."""
    headers = {'apikey': key, 'Authorization': f'Bearer {key}'}
    r = requests.get(f'{url}/rest/v1/cert_types?select=id,code', headers=headers)
    r.raise_for_status()
    return {row['code']: row['id'] for row in r.json()}


def insert_cert(url: str, key: str, record: dict, dry_run: bool) -> bool:
    if dry_run:
        print(f"  [DRY] INSERT: {record['full_name'][:40]:40} | {record['cert_type_code']:18} | issued={record.get('issued_at','?')} | exp={record.get('expires_at','∞')}")
        return True
    headers = {
        'apikey': key,
        'Authorization': f'Bearer {key}',
        'Content-Type': 'application/json',
        'Prefer': 'return=minimal,resolution=ignore-duplicates',
    }
    r = requests.post(f'{url}/rest/v1/employee_certs', headers=headers, json=record)
    if r.status_code in (200, 201):
        return True
    # 409 conflict = duplicate, skip silently
    if r.status_code == 409:
        return True
    print(f"  ERROR {r.status_code}: {r.text[:200]}")
    return False


# ============================================================
# MAIN
# ============================================================

def main():
    parser = argparse.ArgumentParser(description='Import certs from Excel to Supabase')
    parser.add_argument('--dry-run', action='store_true', help='Print without inserting')
    parser.add_argument('--sheet', type=str, help='Import only this sheet')
    args = parser.parse_args()

    # Find Excel file
    excel_path = None
    for candidate in EXCEL_CANDIDATES:
        if candidate.exists():
            excel_path = candidate
            break
    if not excel_path:
        print("ERROR: Excel file not found. Tried:")
        for c in EXCEL_CANDIDATES:
            print(f"  {c}")
        print("\nPlace the file at one of the above paths and retry.")
        sys.exit(1)

    print(f"Loading Excel from: {excel_path}")
    wb = openpyxl.load_workbook(str(excel_path))

    # Connect to Supabase
    url, key = get_supabase_client()
    print(f"Supabase URL: {url}")

    users = fetch_users(url, key)
    print(f"Loaded {len(users)} users from DB")

    cert_types = fetch_cert_types(url, key)
    print(f"Loaded {len(cert_types)} cert types from DB")

    name_idx = build_name_index(users)

    # Collect all records
    all_records = []
    skipped_sheets = []

    for sheet_name in wb.sheetnames:
        if args.sheet and sheet_name != args.sheet:
            continue
        if sheet_name in ('Лист1', 'смены (С, В, Л)'):
            continue

        ws = wb[sheet_name]
        print(f"\nParsing sheet: '{sheet_name}'")

        # Parse records
        if sheet_name == 'Бригадиры':
            records = parse_brigadiry_sheet(ws)
        elif sheet_name == 'ОТ':
            records = parse_ot_sheet(ws)
        elif sheet_name == 'Промбез':
            records = parse_promsbez_sheet(ws)
        elif sheet_name in SHEET_TO_CERT_TYPE:
            cert_code = SHEET_TO_CERT_TYPE[sheet_name]
            records = parse_standard_sheet(ws, cert_code)
        else:
            print(f"  SKIP: no cert_type mapping for this sheet")
            skipped_sheets.append(sheet_name)
            continue

        print(f"  Parsed {len(records)} records")

        # Enrich records with cert_type_id and employee_id
        matched = 0
        unmatched = 0
        missing_cert_type = 0

        for rec in records:
            cert_code = rec.pop('cert_type_code')
            cert_type_id = cert_types.get(cert_code)
            if not cert_type_id:
                print(f"  WARN: cert_type '{cert_code}' not found in DB — run migration 035 first!")
                missing_cert_type += 1
                continue

            rec['cert_type_id'] = cert_type_id

            # Try to match employee
            employee_id = fuzzy_match(rec['full_name'], name_idx)
            if employee_id:
                rec['employee_id'] = employee_id
                matched += 1
            else:
                rec['employee_id'] = None
                unmatched += 1

            # issued_at is required — skip if missing
            if not rec.get('issued_at') and not rec.get('is_indefinite'):
                # Still insert with null issued_at? Or skip?
                # Keep with null — safety engineer can fill manually
                pass

            all_records.append(rec)

        print(f"  Matched: {matched}, Unmatched (no user_id): {unmatched}, Missing cert type: {missing_cert_type}")

    # Insert all records
    print(f"\n{'[DRY RUN] ' if args.dry_run else ''}Inserting {len(all_records)} records...")

    ok = 0
    fail = 0
    unmatched_names = set()

    for rec in all_records:
        if rec.get('employee_id') is None:
            unmatched_names.add(rec.get('full_name', '?'))

        # Build DB payload
        payload = {k: v for k, v in rec.items() if v is not None}
        payload.setdefault('is_indefinite', False)

        if insert_cert(url, key, payload, args.dry_run):
            ok += 1
        else:
            fail += 1

    print(f"\n{'[DRY RUN] ' if args.dry_run else ''}Done: {ok} inserted, {fail} failed")

    if unmatched_names:
        print(f"\n{len(unmatched_names)} names NOT matched to users (saved without employee_id):")
        for n in sorted(unmatched_names):
            print(f"  ⚠ {n}")
        print("\nThese records will appear in the 'Несвязанные' tab in /safety")
        print("Link them manually by opening each record and selecting the employee.")

    if skipped_sheets:
        print(f"\nSkipped sheets (no mapping): {skipped_sheets}")


if __name__ == '__main__':
    main()
